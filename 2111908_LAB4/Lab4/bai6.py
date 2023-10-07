
from openpyxl import *
from tkinter import *
 

wb = load_workbook('Z:\\Lab4\\excel.xlsx')
 

sheet = wb.active
 
 
def excel():
     
    
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50
 
    
    sheet.cell(row=1, column=1).value = "Mã số sinh viên"
    sheet.cell(row=1, column=2).value = "Họ tên"
    sheet.cell(row=1, column=3).value = "Ngày sinhr"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "Số điện thoại"
    sheet.cell(row=1, column=6).value = "Học kỳ"
    sheet.cell(row=1, column=7).value = "Năm học"
 
 

def focus1(event):
    
    mssv_field.focus_set()
 
 

def focus2(event):
    
    hoten_field.focus_set()
 
 

def focus3(event):
    
    ngaysinh_field.focus_set()
 
 

def focus4(event):
    
    email_field.focus_set()
 
 

def focus5(event):
    
    sodt_field.focus_set()
 
 

def focus6(event):
    
    hocky_field.focus_set()
    
def focus7(event):
    
    namhoc_field.focus_set()
 
 

def clear():
     
    
    mssv_field.delete(0, END)
    hoten_field.delete(0, END)
    ngaysinh_field.delete(0, END)
    email_field.delete(0, END)
    sodt_no_field.delete(0, END)
    hocky_field.delete(0, END)
    namhoc_field.delete(0, END)
 
 

def insert():
     
   
    if (mssv_field.get() == "" and
        hoten_field.get() == "" and
        ngaysinh_field.get() == "" and
        email_no_field.get() == "" and
        sodt_no_field.get() == "" and
        hocky_field.get() == "" and
       namhoc_field.get() == ""):
             
        print("empty input")
 
    else:
 
        
        current_row = sheet.max_row
        current_column = sheet.max_column
 
        
        sheet.cell(row=current_row + 1, column=1).value = mssv_field.get()
        sheet.cell(row=current_row + 1, column=2).value = hoten_field.get()
        sheet.cell(row=current_row + 1, column=3).value = ngaysinh_field.get()
        sheet.cell(row=current_row + 1, column=4).value = email_field.get()
        sheet.cell(row=current_row + 1, column=5).value = sodt_field.get()
        sheet.cell(row=current_row + 1, column=6).value = hocky_field.get()
        sheet.cell(row=current_row + 1, column=7).value = namhoc_field.get()
 
        
        wb.save('Z:\\Lab4\\excel.xlsx')
 
        
        mssv_field.focus_set()
 
        
        clear()
 
 

if __name__ == "__main__":
     
    
    root = Tk()
 
   
    root.configure(background='light green')
 
   
    root.title("registration form")
 
    
    root.geometry("500x300")
 
    excel()
 
   
    heading = Label(root, text="THÔNG TIN ĐĂNG KÍ HỌC PHẦN", bg="light green")
 
    
    mssv = Label(root, text="Mã số sinh viên", bg="light green")
 
    
    hoten = Label(root, text="Họ tên", bg="light green")
 
    
    ngaysinh = Label(root, text="Ngày sinh", bg="light green")
 
    email = Label(root, text="Email", bg="light green")
 
    
    sodt = Label(root, text="Số điện thoại", bg="light green")
 

    
 
    
    hocky = Label(root, text="Học kỳ", bg="light green")

    namhoc = Label(root, text="Năm học", bg="light green")
    
    
 
    
    heading.grid(row=0, column=1)
    mssv.grid(row=1, column=0)
    hoten.grid(row=2, column=0)
    ngaysinh.grid(row=3, column=0)
    email.grid(row=4, column=0)
    sodt.grid(row=5, column=0)
    hocky.grid(row=6, column=0)
    namhoc.grid(row=7, column=0)
 
   
    mssv_field = Entry(root)
    hoten_field = Entry(root)
    ngaysinh_field = Entry(root)
    email_field = Entry(root)
    sodt_field = Entry(root)
    hocky_field = Entry(root)
    namhoc_field = Entry(root)
 
   
    mssv_field.bind("<Return>", focus1)
 
    
    hoten_field.bind("<Return>", focus2)
 
  
    ngaysinh_field.bind("<Return>", focus3)
 
    
    email_field.bind("<Return>", focus4)
 
    
    sodt_field.bind("<Return>", focus5)
 
    
    hocky_field.bind("<Return>", focus6)

    namhoc_field.bind("<Return>", focus7)
 
    
    mssv_field.grid(row=1, column=1, ipadx="100")
    hoten_field.grid(row=2, column=1, ipadx="100")
    ngaysinh_field.grid(row=3, column=1, ipadx="100")
    email_field.grid(row=4, column=1, ipadx="100")
    sodt_field.grid(row=5, column=1, ipadx="100")
    hocky_field.grid(row=6, column=1, ipadx="100")
    namhoc_field.grid(row=7, column=1, ipadx="100")
 
   
    excel()
 
    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                            bg="Red", command=insert)
    submit.grid(row=8, column=1)
 
    # start the GUI
    root.mainloop()